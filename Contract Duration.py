from openpyxl import load_workbook
import pandas as pd

# Replace 'Sheet.xlsx' with your actual file path
file_path = 'Sheet.xlsx'

# Read the Excel file
excel_data = pd.read_excel(file_path)

# Load the workbook
wb = load_workbook(file_path)
ws = wb['Sheet1']  # Specify the sheet name

# Iterate through rows where 'Feedback' is 'rented'
for index, row in excel_data[excel_data['Feedback'] == 'rented'].iterrows():
    # Extract the 'Monthly Contract' information
    monthly_contract = str(row['Monthly Contract'])
    
    # Parse the number of months from the contract string
    contract_duration = int(''.join(filter(str.isdigit, monthly_contract)))
    
    # Parse the start date and add the contract duration to calculate the end date
    start_date = pd.to_datetime(row['Start Date'], errors='coerce')
    if not pd.isnull(start_date):
        end_date = start_date + pd.DateOffset(months=contract_duration)
        ws[f'I{index + 2}'] = end_date.strftime('%Y-%m-%d')

# Save the updated workbook
wb.save(file_path)

print(f'End Dates updated in the "End Date" column (column I) in the Excel file: {file_path}')
