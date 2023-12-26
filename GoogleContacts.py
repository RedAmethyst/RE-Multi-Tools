import openpyxl
import re

# Function to clean and format a phone number
def clean_and_format(phone_number):
    phone_number = re.sub(r'\D', '', phone_number)
    if phone_number.startswith('00'):
        return '+971' + phone_number[2:]
    elif phone_number.startswith('0'):
        if len(phone_number) == 10 and phone_number[1] in '5':
            return '+971' + phone_number[1:]
        elif len(phone_number) == 9 and phone_number.startswith('5'):
            return '+971' + phone_number
        else:
            return '+' + phone_number[1:]
    elif phone_number.startswith('971') or phone_number.startswith('+971') or phone_number.startswith('00971'):
        phone_number = re.sub(r'^[+|0]*971', '', phone_number)
        return '+971' + phone_number[-9:]
    elif len(phone_number) == 9 and phone_number.startswith('5'):
        return '+971' + phone_number
    else:
        return '+' + phone_number[1:]

# Dictionary to map country codes to tags
country_tags = {
    '+971': ' UAE',
    '+1': ' USA',
    '+91': ' India',
    '+86': ' China',
    '+44': ' UK',
    '+98': ' Iran',
    '+7': ' Russia',
    '+966': ' KSA',
    '+92': ' PAK',
    '+90': ' TURKY',
    '+49': ' GERMANY',
    '+213': ' DZA',
    '+212': ' MOROCCO',
    '+973': ' BAHRAIN',
    '+965': ' Kuwait',
    '+961': ' Lebanon',
    '+968': ' Oman',
    '+970': ' Palestine',
    '+974': ' Qatar',
    '+20': ' Egypt',
    '+93': ' Afghanistan',
    '+249': ' Sudan',
    '+998': ' Uzbekistan',
}

# Open the Excel file in read-only mode
input_file = 'Sheet.xlsx'  # Change the Excel file name accordingly
wb = openpyxl.load_workbook(input_file, read_only=True)
sheet = wb.active

# Get column indexes for "Phone Number", "Name", and "Sub Location"
phone_number_column = None
name_column = None
sub_location_column = None

for row in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row:
        if cell.value == "Phone Number":
            phone_number_column = cell.column
        elif cell.value == "Name":
            name_column = cell.column
        elif cell.value == "Sub Location":
            sub_location_column = cell.column

if phone_number_column and name_column and sub_location_column:
    # Retrieve the content of cell B2 for output file naming
    output_file_name = str(sheet['B2'].value) + '.vcf'

    # Open a single VCard file for all contacts
    with open(output_file_name, 'w', encoding='utf-8') as vcard_file:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            phone_number = clean_and_format(str(row[phone_number_column - 1].value))
            name = str(row[name_column - 1].value)
            sub_location = str(row[sub_location_column - 1].value)  # Extract sub-location

            # Modify the name based on the phone number format and include sub-location
            country_tag = " Forg"  # Default tag if no country code matches
            for country_code, tag in country_tags.items():
                if phone_number.startswith(country_code):
                    country_tag = tag
                    break
            name_with_location = f"{name} - {sub_location}{country_tag}"

            # Write a VCard entry for each contact to the file
            vcard_file.write(f'BEGIN:VCARD\n')
            vcard_file.write(f'FN:{name_with_location}\n')
            vcard_file.write(f'ADR;TYPE=HOME:;;{sub_location}\n')  # Add sub-location here
            vcard_file.write(f'TEL:{phone_number}\n')
            vcard_file.write(f'END:VCARD\n')

    print(f"VCard file '{output_file_name}' with all contacts has been created.")
else:
    print("Error: Unable to find the required columns in the Excel file.")

# Close the workbook
wb.close()
